from xlsxwriter import Workbook
from openpyxl   import load_workbook

def number_indexes_to_excel_notation(x, y):
    return str(chr( (ord('A') + y) )) + str(x + 1)
def excel_notation_to_number_indexes(excel_note):
    return [int(ord(excel_note[0].upper()) - ord('A') + 1),int(excel_note[1])]

class coordinates:
    def __init__(self, data:list, width:int):
        self.data = data
        self.width = width - 1
        self.excel_range = [None, None]
        self.height = None

    def set_excel_range(self, x, y):
        self.excel_range[0] = number_indexes_to_excel_notation(x, y)
        self.excel_range[1] = number_indexes_to_excel_notation(x + self.height, y + self.width)
    def get_excel_range(self):
        return f"{self.excel_range[0]}:{self.excel_range[1]}"
    def __str__(self):
        return f"data: {self.data}, excel_location: {self.excel_range}, height: {self.height}, width: {self.width}."


class modelos_lineal:
    # constructor de clase: modelos().
    def __init__(self, input_filename, output_filename):
        self.x = coordinates([],1)
        self.y = coordinates([],1)
        self.aprox = coordinates([],1)
        self.error = coordinates([],1)
        self.matrix_A = coordinates([[], []],2)
        self.matrix_Y = coordinates([],1)
        self.input_filename = input_filename
        self.output_filename = output_filename
        self.A_transpuesta_por_A = coordinates([],2)
        self.A_transpuesta_por_Y = coordinates([[], []],1)
        self.c_0 = coordinates(0,1)
        self.c_1 = coordinates(0,1)
        
    
        self.workbook_input = load_workbook(self.input_filename)
        self.worksheet_input = self.workbook_input["input"]
        for row in self.worksheet_input:
            self.x.data.append(row[0].value)
            self.y.data.append(row[1].value)
        self.workbook_input.close()

        self.workbook_output = Workbook(self.output_filename)
        self.worksheet_output = self.workbook_output.add_worksheet("output")
        
        self.init_x()
        self.init_y()
        self.init_matrix_A()
        self.init_matrix_Y()
        self.write_everything()

        self.workbook_output.close()

    def init_x(self):
        self.x.height = len(self.x.data) - 2
    def init_y(self):
        self.y.height = len(self.y.data) - 2
    def init_aprox(self):
        self.aprox.height = len(self.aprox.data) - 2
    def init_error(self):
        self.error.height = len(self.error.data) - 2
    def init_matrix_A(self):
        for i in self.x.data[1:]:
            self.matrix_A.data[0].append(1)
            self.matrix_A.data[1].append(i)
        self.matrix_A.height = min(len(self.matrix_A.data[0]),len(self.matrix_A.data[1])) - 1
    def init_matrix_Y(self):
        for i in self.y.data[1:]:
            self.matrix_Y.data.append(i)
        self.matrix_Y.height = len(self.matrix_Y.data) - 1
            
    def write_everything(self):
        # copy the x, y, 
        i, current_row_pos = 1, 0
        self.x.set_excel_range(i, current_row_pos)
        self.y.set_excel_range(i, current_row_pos + 1)
        self.worksheet_output.write(0, current_row_pos, self.x.data[0])
        self.worksheet_output.write(0, current_row_pos+1, self.y.data[0])
        for x,y in zip(self.x.data[1:], self.y.data[1:]):
            self.worksheet_output.write(i,current_row_pos,x)
            self.worksheet_output.write(i,current_row_pos + 1,y)
            i += 1
        current_row_pos += 2
        self.worksheet_output.write(0, current_row_pos, "Aproximado")
        self.worksheet_output.write(0, current_row_pos + 1, "Error")
        current_row_pos += 3
        i = 1
        self.worksheet_output.write(0,current_row_pos, "A")
        self.matrix_A.set_excel_range(i, current_row_pos)
        for col1, col2 in zip(self.matrix_A.data[0], self.matrix_A.data[1]):
            self.worksheet_output.write(i, current_row_pos, col1)
            self.worksheet_output.write(i, current_row_pos + 1, col2)
            i += 1
        current_row_pos += 2
        i = 1
        self.worksheet_output.write(0,current_row_pos, "Y")
        self.matrix_Y.set_excel_range(i, current_row_pos)
        for col1 in self.matrix_Y.data:
            self.worksheet_output.write(i, current_row_pos, col1)
            i += 1
        i = 1
        current_row_pos += 2
        self.worksheet_output.write(0, current_row_pos, "(A^T)*A")
        self.worksheet_output.write_array_formula(
            number_indexes_to_excel_notation(1, current_row_pos)+":"+number_indexes_to_excel_notation(1+1, current_row_pos+1)
            , "{"+f"=MMULT(TRANSPOSE({self.matrix_A.get_excel_range()}),{self.matrix_A.get_excel_range()})"+"}")
        print(f"=MMULT(TRANSPOSE({self.matrix_A.get_excel_range()}),{self.matrix_A.get_excel_range()})")
        self.A_transpuesta_por_A.excel_range = str(number_indexes_to_excel_notation(1,current_row_pos)) + "#"

        self.worksheet_output.write(5, current_row_pos, "(A^T)*Y")
        self.worksheet_output.write_formula(number_indexes_to_excel_notation(6, current_row_pos), "{"+"=MMULT(TRANSPOSE({}),{})".format(self.matrix_A.get_excel_range(), self.matrix_Y.get_excel_range())+"}")
        self.worksheet_output.write_array_formula(
            number_indexes_to_excel_notation(6, current_row_pos)+":"+number_indexes_to_excel_notation(6+1, current_row_pos),
            "{"+"=MMULT(TRANSPOSE({}),{})".format(self.matrix_A.get_excel_range(), self.matrix_Y.get_excel_range())+"}"
        )
        print(f"=MMULT(TRANSPOSE({self.matrix_A.get_excel_range()}),{self.matrix_Y.get_excel_range()})")
        self.A_transpuesta_por_Y.excel_range = str(number_indexes_to_excel_notation(6,current_row_pos)) + "#"
        print(f"x: {self.x.get_excel_range()}, y: {self.y.get_excel_range()}, matrix_A: {self.matrix_A.get_excel_range()}, matrix_Y: {self.matrix_Y.get_excel_range()}, (A^T)A: {self.A_transpuesta_por_A.excel_range}, (A^T)Y: {self.A_transpuesta_por_Y.excel_range}")

        # self.worksheet_output.write_formula()
        


if __name__ == "__main__":
    modelo = modelos_lineal("./input.xlsx", "./output.xlsx")
